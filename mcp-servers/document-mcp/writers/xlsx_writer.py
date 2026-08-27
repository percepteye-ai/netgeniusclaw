"""Workbooks. Spec 082, US2.

Two things here are not cosmetic.

**Every string cell is forced to inlineStr.** openpyxl writes a leading `=` as a live
formula (measured — research D5), and the strings this writes come from device
descriptions and ticket fields. See sanitize.py.

**Failed devices are rows, not omissions.** A shorter spreadsheet reads as a smaller
estate, which is a false statement about the network. So a device that could not be
reached occupies a visibly marked row and is counted in the banner, and the row count
reflects what was *attempted*.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from guards import apply_bound, max_rows, reject_merged_status, reject_template
from outcomes import (
    Outcome,
    TaggedValue,
    ValueRefused,
    parse_tagged,
    render_as_of,
    render_source,
    render_tagged,
)
from provenance import DocumentStamp, SourceLedger
from sanitize import force_text_cell, write_cell

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="333333")
GAP_FONT = Font(bold=True, color="990000")
BANNER_FONT = Font(bold=True)
BANNER_FILL = PatternFill("solid", fgColor="FFF2CC")
MUTED_FONT = Font(size=8, color="606060")


def build(payload: dict, stamp: DocumentStamp, ledger: SourceLedger) -> tuple[bytes, list[str]]:
    reject_template(payload)

    sheets = payload.get("sheets") or []
    if not isinstance(sheets, list) or not sheets:
        raise ValueRefused("'sheets' must be a non-empty list", Outcome.REFUSED_UNTYPED)

    caveats: list[str] = []
    wb = Workbook()
    wb.remove(wb.active)

    for s_index, sheet in enumerate(sheets):
        if not isinstance(sheet, dict):
            raise ValueRefused(f"sheets[{s_index}] must be an object", Outcome.REFUSED_UNTYPED)
        name = str(sheet.get("name") or f"Sheet{s_index + 1}")[:31]
        columns = [str(c) for c in (sheet.get("columns") or [])]
        if not columns:
            raise ValueRefused(f"sheets[{s_index}]: 'columns' is required", Outcome.REFUSED_UNTYPED)

        # FR-015 — admin state and operational state are different facts.
        reject_merged_status(columns, name)

        rows = list(sheet.get("rows") or [])
        failed_rows = list(sheet.get("failed_rows") or [])
        attempted = len(rows) + len(failed_rows)

        rows, truncated = apply_bound(rows, max_rows())
        if truncated:
            stamp.truncated = True
            stamp.bound_applied = max_rows()
            stamp.bound_kind = "rows per sheet"

        ws = wb.create_sheet(title=name)
        r = 1

        # Banner: attempted / succeeded / failed, frozen so it survives scrolling.
        # SC-004 — the row count must reflect what was attempted, not what worked.
        banner = (
            f"{attempted} attempted · {len(rows)} returned data · {len(failed_rows)} failed. "
            f"{stamp.footer_text()}"
        )
        cell = ws.cell(row=r, column=1)
        force_text_cell(cell, banner)
        cell.font = BANNER_FONT
        cell.fill = BANNER_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(columns) + 2)
        r += 1

        if stamp.truncated:
            tcell = ws.cell(row=r, column=1)
            force_text_cell(tcell, stamp.truncation_text())
            tcell.font = GAP_FONT
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(columns) + 2)
            r += 1

        header_row = r
        # The writer appends Source and As of — a caller cannot omit attribution.
        for c, header in enumerate(columns + ["Source", "As of"], start=1):
            hc = ws.cell(row=header_row, column=c)
            force_text_cell(hc, header)
            hc.font = HEADER_FONT
            hc.fill = HEADER_FILL
        r += 1

        for row_index, raw_row in enumerate(rows):
            tvs: list[TaggedValue] = []
            for c, raw in enumerate(list(raw_row)[: len(columns)]):
                tv = ledger.record(
                    parse_tagged(raw, f"sheets[{s_index}].rows[{row_index}][{c}]")
                )
                tvs.append(tv)
                cell = ws.cell(row=r, column=c + 1)
                if tv.kind == "value" and isinstance(tv.value, (int, float)) and not isinstance(
                    tv.value, bool
                ):
                    write_cell(cell, tv.value)
                else:
                    force_text_cell(cell, render_tagged(tv))
                if tv.is_gap:
                    cell.font = GAP_FONT
            srcs = sorted({render_source(t) for t in tvs if t.kind == "value"})
            ages = sorted({t.as_of for t in tvs if t.kind == "value" and t.as_of})
            force_text_cell(
                ws.cell(row=r, column=len(columns) + 1),
                " / ".join(srcs) if srcs else "— (no data)",
            )
            force_text_cell(
                ws.cell(row=r, column=len(columns) + 2),
                ages[-1] if ages else "(not stated by source)",
            )
            r += 1

        # FR-003 — failed devices are rows, positioned with the data.
        for f_index, failed in enumerate(failed_rows):
            label = str(failed.get("label", f"(unnamed {f_index})"))
            reason = str(failed.get("failed") or failed.get("reason") or "")
            if not reason.strip():
                raise ValueRefused(
                    f"sheets[{s_index}].failed_rows[{f_index}]: a failed row requires a "
                    f"reason. Without one it is a blank row wearing a label",
                    Outcome.REFUSED_UNTYPED,
                )
            ledger.record(parse_tagged({"failed": reason}, f"sheets[{s_index}].failed_rows[{f_index}]"))
            first = ws.cell(row=r, column=1)
            force_text_cell(first, label)
            first.font = GAP_FONT
            marker = ws.cell(row=r, column=2)
            force_text_cell(marker, f"RETRIEVAL FAILED — {reason}")
            marker.font = GAP_FONT
            if len(columns) > 2:
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=len(columns))
            force_text_cell(ws.cell(row=r, column=len(columns) + 1), "— (retrieval failed)")
            force_text_cell(ws.cell(row=r, column=len(columns) + 2), "")
            r += 1

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        _autosize(ws, len(columns) + 2)

        if failed_rows:
            caveats.append(
                f"sheet {name!r}: {len(failed_rows)} device(s) could not be reached and "
                f"appear as failed rows. The sheet is not shorter for it — a shorter "
                f"sheet would read as a smaller estate"
            )

    _add_sources_sheet(wb, ledger, stamp)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), caveats


def _autosize(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        longest = 0
        for cell in ws[get_column_letter(c)]:
            if cell.value is not None:
                longest = max(longest, min(len(str(cell.value)), 60))
        ws.column_dimensions[get_column_letter(c)].width = max(12, longest + 2)


def _add_sources_sheet(wb: Workbook, ledger: SourceLedger, stamp: DocumentStamp) -> None:
    ws = wb.create_sheet(title="Sources")
    stamp_cell = ws.cell(row=1, column=1)
    force_text_cell(stamp_cell, stamp.footer_text())
    stamp_cell.font = BANNER_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SourceLedger.SOURCE_COLUMNS))

    note = ws.cell(row=2, column=1)
    force_text_cell(
        note,
        "Every figure in this workbook came from one of the following. This sheet is in "
        "addition to the per-row Source column, not instead of it.",
    )
    note.font = MUTED_FONT
    note.alignment = Alignment(wrap_text=False)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(SourceLedger.SOURCE_COLUMNS))

    for c, header in enumerate(SourceLedger.SOURCE_COLUMNS, start=1):
        hc = ws.cell(row=3, column=c)
        force_text_cell(hc, header)
        hc.font = HEADER_FONT
        hc.fill = HEADER_FILL

    r = 4
    for row in ledger.as_rows():
        for c, value in enumerate(row, start=1):
            force_text_cell(ws.cell(row=r, column=c), value)
        r += 1
    _autosize(ws, len(SourceLedger.SOURCE_COLUMNS))
