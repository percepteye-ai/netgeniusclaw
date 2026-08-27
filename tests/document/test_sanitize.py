"""Untrusted text. Spec 082, FR-026, SC-015.

The raw-XML assertion below is the point of this suite. `cell.data_type == "s"` is an
openpyxl-level claim; `<f>` not appearing in xl/worksheets/sheet1.xml is a claim about
the file an auditor actually opens. Measured 2026-08-03, a naive write produces
`<c r="A1"><f>1+1</f><v></v></c>` — a live formula built from a device description.
"""

from __future__ import annotations

import re
import zipfile

from _harness import FAILURES, abspath, check, cleanup, run, sandbox  # noqa: F401

import output  # noqa: E402
from provenance import DocumentStamp, SourceLedger  # noqa: E402
from sanitize import plain_text  # noqa: E402
from writers import xlsx_writer  # noqa: E402

HOSTILE = [
    "=1+1",
    "=cmd|'/c calc'!A0",
    "@SUM(1,1)",
    "+1+1",
    "-1+1",
    "=HYPERLINK(\"http://evil\",\"click\")",
]


def _workbook_with(values):
    d = sandbox()
    stamp = DocumentStamp(tool="xlsx_write")
    ledger = SourceLedger()
    payload = {
        "sheets": [
            {
                "name": "Data",
                "columns": ["Interface", "Description"],
                "rows": [
                    [{"v": f"port{i}", "src": "fgt_list_interfaces"},
                     {"v": v, "src": "fgt_list_interfaces"}]
                    for i, v in enumerate(values)
                ],
            }
        ]
    }
    data, _ = xlsx_writer.build(payload, stamp, ledger)
    path, suffix = output.reserve("xlsx", "sanitize")
    path.write_bytes(data)
    return d, path


def test_no_formula_survives_into_the_file():
    d, path = _workbook_with(HOSTILE)
    try:
        with zipfile.ZipFile(str(path)) as z:
            names = [n for n in z.namelist() if n.startswith("xl/worksheets/")]
            xml = "".join(z.read(n).decode() for n in names)
        check(
            "no <f> element anywhere in the worksheet XML",
            "<f>" not in xml and "<f " not in xml,
            "a formula element is present — untrusted text became executable content",
        )
        check(
            "hostile strings are stored as inline text",
            xml.count("inlineStr") >= len(HOSTILE),
            f"only {xml.count('inlineStr')} inlineStr cells for {len(HOSTILE)} hostile values",
        )
    finally:
        cleanup(d)


def test_values_round_trip_uncorrupted():
    """The common mitigation is an apostrophe prefix. It visibly corrupts the value, in
    a document whose entire purpose is fidelity. This asserts we did not do that."""
    d, path = _workbook_with(HOSTILE)
    try:
        import openpyxl

        ws = openpyxl.load_workbook(str(path))["Data"]
        header_row = None
        for r in range(1, 8):
            if ws.cell(row=r, column=1).value == "Interface":
                header_row = r
                break
        check("header row was found", header_row is not None)
        if header_row:
            for i, expected in enumerate(HOSTILE):
                cell = ws.cell(row=header_row + 1 + i, column=2)
                check(
                    f"{expected!r} round-trips exactly",
                    cell.value == expected,
                    f"got {cell.value!r} — the mitigation corrupted the text",
                )
                check(
                    f"{expected!r} is typed as a string",
                    cell.data_type == "s",
                    f"data_type={cell.data_type!r}",
                )
    finally:
        cleanup(d)


def test_numbers_stay_numbers():
    """Forcing everything to text would break a spreadsheet's ability to sum a column,
    which is most of why an auditor asked for .xlsx and not .csv."""
    d = sandbox()
    try:
        stamp, ledger = DocumentStamp(tool="xlsx_write"), SourceLedger()
        payload = {"sheets": [{"name": "N", "columns": ["Count"],
                               "rows": [[{"v": 42, "src": "x"}], [{"v": 3.5, "src": "x"}]]}]}
        data, _ = xlsx_writer.build(payload, stamp, ledger)
        path, _s = output.reserve("xlsx", "numbers")
        path.write_bytes(data)
        import openpyxl

        ws = openpyxl.load_workbook(str(path))["N"]
        found = [ws.cell(row=r, column=1).value for r in range(1, 8)]
        check("integer stayed numeric", 42 in found, str(found))
        check("float stayed numeric", 3.5 in found, str(found))
    finally:
        cleanup(d)


def test_control_characters_are_stripped():
    """OOXML rejects most control characters; a document that will not open is a worse
    outcome than one with a slightly shortened description."""
    dirty = "port1\x00\x07 desc\x1f"
    clean = plain_text(dirty)
    check("NUL is stripped", "\x00" not in clean, repr(clean))
    check("BEL is stripped", "\x07" not in clean, repr(clean))
    check("printable text survives", "port1" in clean and "desc" in clean, repr(clean))
    check("newlines and tabs survive", plain_text("a\nb\tc") == "a\nb\tc", repr(plain_text("a\nb\tc")))


TESTS = [
    test_no_formula_survives_into_the_file,
    test_values_round_trip_uncorrupted,
    test_numbers_stay_numbers,
    test_control_characters_are_stripped,
]

if __name__ == "__main__":
    raise SystemExit(run(TESTS, "sanitize"))
